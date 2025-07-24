from io import BytesIO
from typing import List

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.file_converters.base_strategy import ReportFormatStrategy
from app.models.report import BoilerRoomEngineerReport


class XLSXStrategy(ReportFormatStrategy):
    def _convert(self, data: List["BoilerRoomEngineerReport"], year: int, **kwargs) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет инженера котельной"

        # Заголовки: "Показатель", "Год", "1", "2", ..., "31"
        headers = ["Показатель", "Год"] + [str(day) for day in range(1, 32)]
        ws.append(headers)

        # Применение стилей к заголовкам
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill

        # Список показателей и соответствующих полей модели
        indicators = [
            ("Отпуск тепловой энергии 8_7 Гкал", "heat_energy_supply_8_to_7_gcal"),
            ("Тепловычислитель магистрали выхода из котельной", "heat_calculator_main_exit_boiler_room"),
            ("Температура в реке Плюсса", "temperature_river_plussa"),
        ]

        archive_indicators = [
            ("Отпуск тепловой энергии 8_7 Гкал (архив)", "archive_heat_energy_supply_8_to_7_gcal"),
            (
                "Тепловычислитель магистрали выхода из котельной (архив)",
                "archive_heat_calculator_main_exit_boiler_room",
            ),
            ("Температура в реке Плюсса (архив)", "archive_temperature_river_plussa"),
        ]

        # Функция для добавления строки данных
        def add_data_row(indicator_name: str, field_name: str, report_year: int):
            row = [indicator_name, report_year]
            for day in range(1, 32):
                # Найти запись для текущего дня
                record = next((item for item in data if item.date.day == day), None)
                if record:
                    value = getattr(record, field_name)
                    row.append(float(value) if value is not None else None)
                else:
                    row.append(None)
            ws.append(row)

        # Добавление строк для текущего года
        for indicator_name, field_name in indicators:
            add_data_row(indicator_name, field_name, year)

        # Добавление строк для архивного года (year - 1)
        for indicator_name, field_name in archive_indicators:
            add_data_row(indicator_name, field_name, year - 1)

        # Применение стилей к данным и форматирование
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column, max_row=ws.max_row):
            # Выравнивание ячеек
            for cell in row:
                if cell.column == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Форматирование чисел
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            # Применение фона для архивных данных
            if row[1].value == year - 1:
                for cell in row:
                    cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")

        # Автоматическая настройка ширины столбцов
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value)) if cell.value is not None else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Сохранение в буфер
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _get_response(self, data: BytesIO, filename: str = "report.xlsx") -> StreamingResponse:
        return StreamingResponse(
            data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
